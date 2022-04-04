import pyodbc
import table_control as tableControl
import datetime
import full_text_search as fullSearch
import random
from openpyxl import Workbook,load_workbook
import base64

class Database_insert:

    def __init__(self):
        self.db = pyodbc.connect(
                            'Driver={ODBC Driver 17 for SQL Server};'
                            'Server=sql.athena.domainhizmetleri.com;'
                            'Database=abdullah_pys;'
                            'UID=abdullah_pys;'
                            'PWD=@PassWord123;'
                            )
    def test(self):
        imlec = self.db.cursor()
        imlec.execute('SELECT * FROM [abdullah_pys].[m_Student]')
        kullanicilar = imlec.fetchall()
        for i in kullanicilar:
            print(i)

    def student_insert(self,student_id,advisior_id,name,surname,mail,
                       phone_no,depart_id,faculty_id,clas,photo_path,password):

        try:
            curs = self.db.cursor()
            curs.execute(
                "insert into m_Student(studentID,advisorID,name,surname,mail,phoneNumber,departmentID,facultyID,class,photoPath,password) values (?, ?,?,?,?,?,?,?,?,?,?)",
                str(student_id),
                int(advisior_id),
                str(name),
                str(surname),
                str(mail),
                str(phone_no),
                str(depart_id),
                str(faculty_id),
                str(clas),
                str(photo_path),
                str(password))
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def faculty_insert(self,faculty_id,name):
        table_c = tableControl.TableControl()
        if not table_c.faculty_id_control(faculty_id):
            if len(name) < 50:
                try:
                    curs = self.db.cursor()
                    curs.execute("insert into m_Faculty(facultyID, name) values (?, ?)", str(faculty_id), str(name))
                    curs.commit()
                    return "Successful"
                except Exception as e:
                    e = str(e)
                    return e
            else:
                return "Error code : 13"
        else:
            return "Error code : 12"

    def department_insert(self,depart_id,faculty_id,name):
        table_c = tableControl.TableControl()
        if not table_c.department_id_control(depart_id):
            if not table_c.faculty_id_control(faculty_id):
                if len(name) < 55:
                    try:
                        curs = self.db.cursor()
                        curs.execute("insert into m_Department(departmentID, facultyID, name) values (?, ?, ?)",
                                     str(depart_id), str(faculty_id), str(name))
                        curs.commit()
                        return "Successful"
                    except Exception as e:
                        e = str(e)
                        return e
                else:
                    return "Error code : 16"
            else:
                return "Error code : 15"
        else:
            return "Error code : 14"

    def advisor_insert(self,name,surname,title,mail,depart_id,faculty_id,photo_path,password):
        try:
            curs = self.db.cursor()
            curs.execute("insert into m_Advisor(name,surname,title,mail,departmentID,facultyID,photoPath,password) values (?,?,?,?,?,?,?,?)",
                str(name),
                str(surname),
                str(title),
                str(mail),
                str(depart_id),
                str(faculty_id),
                str(photo_path),
                str(password))
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def messsage_insert(self,advisor_id,student_id,status,message):
        table_c = tableControl.TableControl()
        date = datetime.datetime.now()
        status = int(status)
        if table_c.advisor_reg_control(advisor_id):
            return "Error code : 26"
        if table_c.student_id_control(student_id):
            return "Error code : 27"
        try:
            curs = self.db.cursor()
            curs.execute("insert into t_message(advisorID,studentID,date,status,message) values (?,?,?,?,?)",
                         int(advisor_id),
                         str(student_id),
                         date,
                         int(status),
                         str(message))
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def dissertation_insert(self,projenumber,pdfpath,docpath,status,desc,insertdate,updatedate):

        try:
            curs = self.db.cursor()
            curs.execute(
                "insert into t_Dissertation(projectNumber,pdfPath,docPath,status,description,insertionDate,updateDate) values (?,?,?,?,?,?,?)",
                projenumber,
                pdfpath,
                docpath,
                status,
                desc,
                insertdate,
                updatedate)
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def reports_insert(self,projenumber,pdfpath,docpath,status,desc,insertdate,updatedate):
        try:
            curs = self.db.cursor()
            curs.execute(
                "insert into t_reports(projectNumber,pdfPath,docPath,status,description,insertionDate,updatedDate) values (?,?,?,?,?,?,?)",
                projenumber,
                pdfpath,
                docpath,
                status,
                desc,
                insertdate,
                updatedate)
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def plagiarism_insert(self,mainprojeid,otherprojeid,plagrismrate):
        try:
            curs = self.db.cursor()
            curs.execute("insert into t_Plagiarism(mainProjeID,otherProjeID,plagiarismRate) values (?,?,?)",
                         mainprojeid,
                         otherprojeid,
                         plagrismrate)
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def semester_insert(self,startdate,enddate,name):
        try:
            curs = self.db.cursor()
            curs.execute("insert into m_semester(startDate,endDate,name) values (?,?,?)",
                         startdate,
                         enddate,
                         name)
            curs.commit()
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def projects_insert(self,headline,matter,cont,purpose,keyword,
                        metariel,method,poss,status,descr,semeterid,
                        studentid,insertiondate,updatedate):
        try:
            version = 1
            deg = random.randint(1000,9999)
            number = str(studentid[:8]) + str(deg)
            curs = self.db.cursor()
            curs.execute("insert into t_Projects(number,version,headline,matter,[content],purpose,keyword,materiel,method,possibility,status,description,semesterID,studentID,insertionDate,updatedDate) values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            number,
                            version,
                            headline,
                             matter,
                             cont,
                             purpose,
                             keyword,
                             metariel,
                             method,
                             poss,
                             status,
                             descr,
                             semeterid,
                             studentid,
                             insertiondate,
                             updatedate)
            curs.commit()
            search = fullSearch.Search()
            search.allin(number)
            return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def advisor_xlsx_add(self,rowCount):
        kisiler = []
        try:
            wb = load_workbook("excel/advisor.xlsx")
            ws = wb.active
            for i in range(2,rowCount+1):
                name = ws["B"+str(i)].value
                surname = ws["C"+str(i)].value
                title = ws["D"+str(i)].value
                mail = ws["E"+str(i)].value
                depart = ws["F"+str(i)].value
                faculty = ws["G"+str(i)].value
                photo = ws["H"+str(i)].value
                passwd = ws["I"+str(i)].value
                kisiler.append([name,surname,title,mail,depart,faculty,photo,passwd])
            return kisiler
        except Exception as e:
            e = str(e)
            return e

    def adv_list_import(self,Rcount,base):
        try:
            with open("excel/advisor.xlsx", "wb") as fileOpen:
                decode64 = base64.b64decode(base)
                fileOpen.write(decode64)
            listem = self.advisor_xlsx_add(int(Rcount))
            if type(listem) == list:
                for a in listem:
                    self.advisor_insert(a[0],a[1],a[2],a[3],a[4],a[5],a[6],a[7])
                return "Successful"
        except Exception as e:
            e = str(e)
            return e

    def student_xlsx_add(self,rowCount):
        kisiler = []
        try:
            wb = load_workbook("excel/student.xlsx")
            ws = wb.active
            for i in range(2,rowCount+1):
                studentid = ws["A"+str(i)].value
                advisorid = ws["B"+str(i)].value
                name = ws["C"+str(i)].value
                surname = ws["D"+str(i)].value
                mail = ws["E"+str(i)].value
                phone = ws["F"+str(i)].value
                depart = ws["G"+str(i)].value
                faculty = ws["H"+str(i)].value
                clas = ws["I"+str(i)].value
                photopath = ws["J"+str(i)].value
                passw = ws["K"+str(i)].value
                kisiler.append([studentid,advisorid,name,surname,mail,phone,depart,faculty,clas,photopath,passw])
            return kisiler
        except Exception as e:
            e = str(e)
            return e

    def student_list_import(self,Rcount,base):
        try:
            with open("excel/student.xlsx", "wb") as fileOpen:
                decode64 = base64.b64decode(base)
                fileOpen.write(decode64)
            listem = self.student_xlsx_add(int(Rcount))

            if type(listem) == list:
                for a in listem:
                    self.student_insert(a[0],a[1],a[2],a[3],a[4],a[5],a[6],a[7],a[8],a[9],a[10])

                return "Successful"
            else:
                return "liste değil"
        except Exception as e:
            e = str(e)
            return e

