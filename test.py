from smtplib import SMTP

# subject = "Konu"
# message = "webmail.abdullahaligun.com -465"
# content = "Subject: {0} \n {1}".format(subject,message)
#
# myMailAdress = "mail@pys.abdullahaligun.com"
# password = "xa1@6C4x"
# # Kime Gönderilecek Bilgisi
# sendTo = "hacklojen01@gmail.com"
#
# mail = SMTP("webmail.abdullahaligun.com",465)
# mail.ehlo()
# mail.starttls()
# mail.login(myMailAdress, password)
# mail.sendmail(sendTo, sendTo, content.encode("utf-8"))


# import smtplib
#
# content = "merhaba"
# mail = smtplib.SMTP("smtp.gmail.com",587)
# mail.ehlo()
# mail.starttls()
# mail.login("info.kouai@gmail.com","")
# mail.sendmail("info.kouai@gmail.com","hacklojen01@gmail.com",content)

from openpyxl import Workbook,load_workbook

wb = load_workbook("documantation/test.xlsx")
ws = wb.active
satir = 5
for i in range(2,satir+1):
    registid = ws["A"+str(i)].value
    name = ws["B"+str(i)].value
    surname = ws["C"+str(i)].value
    title = ws["D"+str(i)].value
    mail = ws["E"+str(i)].value
    depart = ws["F"+str(i)].value
    faculty = ws["G"+str(i)].value
    photo = ws["H"+str(i)].value
    passwd = ws["I"+str(i)].value
    print(str(registid) + name + surname + title + mail + str(depart) + str(faculty) + photo + passwd)


